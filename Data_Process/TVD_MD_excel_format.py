# !/usr/bin/python
# -*- coding:gbk -*-

from openpyxl import Workbook
from openpyxl import load_workbook


class excel_process:
    def __init__(self, inputfile,outputpath):
        self.filename = inputfile
        self.rel_file = ''
        self.new_col = 1
        self.wb = load_workbook(inputfile)
        self.ws_old = self.wb.active
        self.ws_new = self.wb.create_sheet('New')
        self.FormatOLDExcel()
        self.FormatNEWExcel()
        self.DataCopy()
        self.addWellid()
        self.wb.remove_sheet(self.ws_old)
        if  'MD' in self.filename:
            outputfile=outputpath+self.wellid+'MD.xlsx'
        elif 'TVD' in self.filename:
            outputfile=outputpath+self.wellid+'TVD.xlsx'

        self.wb.save(outputfile)
        self.cleanData(outputfile)
        self.TranToCsv(outputfile)
        #print(self.wellid)
        #print(self.rel_file)

    def Dictionary(self):
        adict = {'层号': 'Number', '层位': 'Abbrev_Formation', '岩性': 'Lithology', '起始深度': 'Top',
                      '终止深度':'Bottom',
                      '有效厚度': 'Thickness', '累积厚度': 'Cum_Thickness', '深电阻率': 'RT',
                      '声波时差': 'AC', '密度': 'DEN', '补偿中子': 'CNL', '泥质含量': 'VCL', '孔隙度': 'PHI',
                      '渗透率': 'PERM', '含气饱和度': 'SOG', '解释结论': 'Result','备注': 'Comments',
                      '砂岩':'Sand','干层':'TIGHT','气层':'GAS','差气层':'POOR GAS','水层':'WATER','含水气层':'GAS&WATER',
                      '气水同层':'GAS&WATER','特征数值':'VALUE_ATTRI','含气水层':'GAS&WATER'
                 }

        return adict

    def SplitColumn(self):
        Split_col=['RT', 'AC', 'DEN', 'CNL', 'VCL', 'PHI', 'PERM', 'SOG']
        return Split_col

    def formatCell(self, inputstring):
        # Delete White Space, Enter from Cell
        astring = inputstring
        if astring is not None:
         if '\n' in astring or '_x000D_' in astring or ' ' in astring:
            astring = astring.replace('\n', '')
            astring = astring.replace('_x000D_', '')
            astring = astring.replace(' ', '')
        return astring

    def FormatOLDExcel(self):
        # Count how many columns are included in this excel file
        old_column = 0
        for column in self.ws_old.columns:
            old_column += 1
            cell_old=self.ws_old.cell(row=1, column=old_column).value
        # Format first row from chinese to english, character transformation referenced by Dictionary function
            for j in range(old_column):
                cell_value=self.formatCell(self.ws_old.cell(row=1, column=j+1).value)
                self.ws_old.cell(row=1, column=j+1).value = cell_value
                if cell_value.encode('gbk') in self.Dictionary().keys():
                    self.ws_old.cell(row=1, column=j+1).value = self.Dictionary()[cell_value.encode('gbk')]
                    cell_value = self.ws_old.cell(row=1, column=j+1).value


    def FormatNEWExcel(self):
        # Copy first row from 成果表 sheet to New sheet,and re-format the column which should be split by AVG,MAX and Min
        new_col=0
        col_old=0
        for column in self.ws_old.columns:
            col_old+=1
            cell_old=self.ws_old.cell(row=1, column=col_old).value

            #the columns which listed in the SplitColumn function will be split.
            if cell_old not in self.SplitColumn():
                new_col+=1
                self.ws_new.cell(row=1, column=new_col).value=cell_old
            else:
                for col in [cell_old+'_MIN', cell_old+'_MAX',cell_old+'_AVG']:
                    new_col+=1
                    self.ws_new.cell(row=1, column=new_col).value = col


    def DataCopy(self):
        # Extract data from old sheet and copy data to new sheet, meanwhile, split data which was integrated in one cell
        # to adapt cell in new sheet reference the SplitColumn function
        old_col = 0
        i = 0
        for x in self.ws_old.rows:
            i += 1
        print(i)

        # identify the column by first cell to decide whether this column belong to column which need to be split
        for cols in self.ws_old.columns:
            old_col += 1
            cell_value_1row = self.ws_old.cell(row= 1, column= old_col).value #query 1st row for each column
            cell_value_1row = cell_value_1row.encode('utf8')
            #print(cell_value_1row)
            if cell_value_1row not in self.SplitColumn():
                old_row = 3
                new_row = 2
                for rows in range(i):

                    cell_value = self.ws_old.cell(row=old_row, column=old_col).value
                    print('the cell_value is %s' % cell_value)

                    if isinstance(cell_value, unicode):
                        if cell_value.encode('gbk') in self.Dictionary().keys():
                            self.ws_new.cell(row=new_row, column=self.new_col).value = self.Dictionary()[cell_value.encode('gbk')]
                            new_row += 1
                            old_row += 1
                    else:
                         self.ws_new.cell(row=new_row, column=self.new_col).value=cell_value
                         new_row += 1
                         old_row += 1
                self.new_col += 1

            else:
                old_row = 3
                new_row = 2
                for rows in range(i):
                    new_col=self.new_col
                    cell_value = self.ws_old.cell(row=old_row, column=old_col).value
                    if cell_value is not None:
                        data_min = self.dataProcess(old_row, old_col)[0]
                        self.ws_new.cell(row=new_row, column=new_col).value = data_min
                        new_col += 1
                        data_max = self.dataProcess(old_row, old_col)[1]
                        self.ws_new.cell(row=new_row, column=new_col).value = data_max
                        new_col += 1
                        data_avg = self.dataProcess(old_row, old_col)[2]
                        self.ws_new.cell(row=new_row, column=new_col).value = data_avg
                        new_col += 1
                    old_row += 1
                    new_row += 1
                self.new_col += 3
        print(self.new_col)



    def dataProcess(self, row, column):
        # Old Sheet data format  which data in cell need to be split
        cell_value = self.ws_old.cell(row=row, column=column).value
        cell_value = cell_value.encode('utf8').replace('_x000D_', '').split('\n')
        return cell_value

    def addWellid(self):
        # Add wellid in the excel file at last column
        filename = self.filename.split('\\')[-1]
        self.wellid = filename[0:filename.find('_')]

        self.ws_new.cell(row=1, column=self.new_col).value = 'wellid'
        new_row = 2
        for rows in self.ws_new.rows:
            self.ws_new.cell(row=new_row, column=self.new_col).value =self.wellid
            print(self.ws_new.cell(row=new_row, column=self.new_col).value)
            new_row += 1


    def TranToCsv(self,inputfile):
        import csv
        wb=load_workbook(inputfile)
        ws=wb.active
        self.rel_file=inputfile[:inputfile.find('.xlsx')]+'.csv'
        csv_file=open(self.rel_file, 'wb')
        cformat=csv.writer(csv_file)
        for row in ws.rows:
            cformat.writerow([cell.value for cell in row])


    def cleanData(self, filename):
        wb = load_workbook(filename)
        ws = wb.active
        row_count = 1
        row_check = 1
        for rows in ws.rows:
            row_count += 1
            while ws.cell(row=row_check,column=1).value is not None:
                row_check += 1
        print(row_check,row_count)
        for rows in range(row_check,row_count):
            print(rows)
            ws.cell(row=rows, column=self.new_col).value = None
        wb.save(filename)






if __name__ == '__main__':
    afile = r'W:\GSR\050000-GeoInformation\052300-Projects\Geoscience database\Input files\Markers\SN0029-12_Result_Table_MD_11162015.xlsx'
    bfile = r'W:\GSR\050000-GeoInformation\052300-Projects\Geoscience database\Input files\Markers\\'
    atest = excel_process(afile,bfile)
